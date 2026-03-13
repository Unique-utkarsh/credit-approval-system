"""
Management command to ingest customer and loan data from Excel files.
Can be run directly (synchronous) or dispatch Celery tasks.
"""
import logging
import datetime
import os

from django.core.management.base import BaseCommand
from django.conf import settings

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Ingest customer and loan data from Excel files into the database'

    def add_arguments(self, parser):
        parser.add_argument(
            '--async',
            action='store_true',
            dest='use_celery',
            help='Dispatch ingestion as Celery background tasks',
        )

    def handle(self, *args, **options):
        if options['use_celery']:
            self._dispatch_celery_tasks()
        else:
            self._run_sync()

    def _dispatch_celery_tasks(self):
        from credit_app.tasks import ingest_customer_data, ingest_loan_data
        self.stdout.write('Dispatching background ingestion tasks via Celery...')
        ingest_customer_data.delay()
        ingest_loan_data.delay()
        self.stdout.write(self.style.SUCCESS('Tasks dispatched successfully.'))

    def _run_sync(self):
        """Run ingestion synchronously — used at container startup."""
        self.stdout.write('Starting synchronous data ingestion...')
        self._ingest_customers()
        self._ingest_loans()
        self.stdout.write(self.style.SUCCESS('Data ingestion complete.'))

    def _ingest_customers(self):
        import openpyxl
        from credit_app.models import Customer

        file_path = str(settings.DATA_DIR / 'customer_data.xlsx')
        if not os.path.exists(file_path):
            self.stderr.write(f'Customer data file not found: {file_path}')
            return

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        created = updated = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            customer_id, first_name, last_name, age, phone_number, monthly_salary, approved_limit = row
            if customer_id is None:
                continue

            _, was_created = Customer.objects.update_or_create(
                customer_id=int(customer_id),
                defaults={
                    'first_name': str(first_name or ''),
                    'last_name': str(last_name or ''),
                    'age': int(age or 0),
                    'phone_number': int(phone_number or 0),
                    'monthly_salary': int(monthly_salary or 0),
                    'approved_limit': int(approved_limit or 0),
                    'current_debt': 0,
                }
            )
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(f'  Customers — created: {created}, updated: {updated}')

    def _ingest_loans(self):
        import openpyxl
        from credit_app.models import Customer, Loan

        file_path = str(settings.DATA_DIR / 'loan_data.xlsx')
        if not os.path.exists(file_path):
            self.stderr.write(f'Loan data file not found: {file_path}')
            return

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        created = skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            (customer_id, loan_id, loan_amount, tenure,
             interest_rate, monthly_payment, emis_paid_on_time,
             start_date, end_date) = row

            if loan_id is None or customer_id is None:
                continue

            try:
                customer = Customer.objects.get(customer_id=int(customer_id))
            except Customer.DoesNotExist:
                skipped += 1
                continue

            if isinstance(start_date, datetime.datetime):
                start_date = start_date.date()
            if isinstance(end_date, datetime.datetime):
                end_date = end_date.date()

            Loan.objects.update_or_create(
                loan_id=int(loan_id),
                defaults={
                    'customer': customer,
                    'loan_amount': float(loan_amount or 0),
                    'tenure': int(tenure or 0),
                    'interest_rate': float(interest_rate or 0),
                    'monthly_repayment': float(monthly_payment or 0),
                    'emis_paid_on_time': int(emis_paid_on_time or 0),
                    'start_date': start_date,
                    'end_date': end_date,
                }
            )
            created += 1

        self.stdout.write(f'  Loans     — processed: {created}, skipped (missing customer): {skipped}')
