import logging
from celery import shared_task
from django.conf import settings

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def ingest_customer_data(self, file_path=None):
    """Background task to ingest customer data from Excel file."""
    import openpyxl
    from credit_app.models import Customer

    if file_path is None:
        file_path = str(settings.DATA_DIR / 'customer_data.xlsx')

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        customers_created = 0
        customers_updated = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            customer_id, first_name, last_name, age, phone_number, monthly_salary, approved_limit = row

            if customer_id is None:
                continue

            _, created = Customer.objects.update_or_create(
                customer_id=int(customer_id),
                defaults={
                    'first_name': str(first_name) if first_name else '',
                    'last_name': str(last_name) if last_name else '',
                    'age': int(age) if age else 0,
                    'phone_number': int(phone_number) if phone_number else 0,
                    'monthly_salary': int(monthly_salary) if monthly_salary else 0,
                    'approved_limit': int(approved_limit) if approved_limit else 0,
                    'current_debt': 0,
                }
            )
            if created:
                customers_created += 1
            else:
                customers_updated += 1

        logger.info(
            f"Customer ingestion complete: {customers_created} created, {customers_updated} updated"
        )
        return {'created': customers_created, 'updated': customers_updated}

    except Exception as exc:
        logger.error(f"Customer ingestion failed: {exc}")
        raise self.retry(exc=exc, countdown=10)


@shared_task(bind=True, max_retries=3)
def ingest_loan_data(self, file_path=None):
    """Background task to ingest loan data from Excel file."""
    import openpyxl
    from credit_app.models import Customer, Loan
    import datetime

    if file_path is None:
        file_path = str(settings.DATA_DIR / 'loan_data.xlsx')

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        loans_created = 0
        loans_skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            (customer_id, loan_id, loan_amount, tenure,
             interest_rate, monthly_payment, emis_paid_on_time,
             start_date, end_date) = row

            if loan_id is None or customer_id is None:
                continue

            try:
                customer = Customer.objects.get(customer_id=int(customer_id))
            except Customer.DoesNotExist:
                logger.warning(f"Customer {customer_id} not found for loan {loan_id}")
                loans_skipped += 1
                continue

            # Handle date parsing
            if isinstance(start_date, datetime.datetime):
                start_date = start_date.date()
            if isinstance(end_date, datetime.datetime):
                end_date = end_date.date()

            _, created = Loan.objects.update_or_create(
                loan_id=int(loan_id),
                defaults={
                    'customer': customer,
                    'loan_amount': float(loan_amount) if loan_amount else 0,
                    'tenure': int(tenure) if tenure else 0,
                    'interest_rate': float(interest_rate) if interest_rate else 0,
                    'monthly_repayment': float(monthly_payment) if monthly_payment else 0,
                    'emis_paid_on_time': int(emis_paid_on_time) if emis_paid_on_time else 0,
                    'start_date': start_date,
                    'end_date': end_date,
                }
            )
            if created:
                loans_created += 1

        logger.info(
            f"Loan ingestion complete: {loans_created} created, {loans_skipped} skipped"
        )
        return {'created': loans_created, 'skipped': loans_skipped}

    except Exception as exc:
        logger.error(f"Loan ingestion failed: {exc}")
        raise self.retry(exc=exc, countdown=10)
